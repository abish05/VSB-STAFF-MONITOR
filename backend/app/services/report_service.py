"""
Report Service — PDF (WeasyPrint + Jinja2) and Excel (openpyxl) generation
"""

import io
import logging
from datetime import datetime, timezone
from typing import Optional

from app.config import settings
from jinja2 import BaseLoader, Environment
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

logger = logging.getLogger(__name__)

# ─── PDF HTML Template ────────────────────────────────────────────────────────
PDF_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<style>
  @page { size: A4; margin: 2cm; }
  body { font-family: Arial, sans-serif; color: #1a1a2e; font-size: 12px; }
  .header { background: linear-gradient(135deg, #6366F1, #22D3EE); color: white; padding: 20px; border-radius: 8px; margin-bottom: 20px; }
  .header h1 { margin: 0; font-size: 22px; }
  .header p { margin: 4px 0; font-size: 11px; opacity: 0.9; }
  .section { margin-bottom: 20px; border: 1px solid #e2e8f0; border-radius: 8px; overflow: hidden; }
  .section-title { background: #6366F1; color: white; padding: 8px 15px; font-weight: bold; font-size: 13px; }
  .section-body { padding: 15px; }
  .stat-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 10px; margin-bottom: 10px; }
  .stat-box { background: #f8fafc; border-radius: 6px; padding: 10px; text-align: center; }
  .stat-box .value { font-size: 20px; font-weight: bold; color: #6366F1; }
  .stat-box .label { font-size: 10px; color: #64748b; margin-top: 3px; }
  .score-bar { background: #e2e8f0; border-radius: 50px; height: 12px; margin: 5px 0; }
  .score-fill { height: 12px; border-radius: 50px; background: linear-gradient(90deg, #6366F1, #22D3EE); }
  .badge { display: inline-block; background: #f1f5f9; border-radius: 20px; padding: 4px 12px; margin: 3px; font-size: 11px; }
  .badge.earned { background: #6366F1; color: white; }
  .ai-section { background: #f8fafc; padding: 12px; border-radius: 6px; margin-top: 8px; }
  .ai-item { margin: 5px 0; padding: 5px 10px; background: white; border-left: 3px solid #6366F1; border-radius: 3px; }
  table { width: 100%; border-collapse: collapse; font-size: 11px; }
  th { background: #6366F1; color: white; padding: 6px 10px; text-align: left; }
  td { padding: 6px 10px; border-bottom: 1px solid #e2e8f0; }
  tr:nth-child(even) td { background: #f8fafc; }
  .footer { text-align: center; font-size: 10px; color: #94a3b8; margin-top: 20px; }
</style>
</head>
<body>

<div class="header">
  <h1>{{ college_name }}</h1>
  <p>CodePulse AI — Performance Report</p>
  <p>Generated: {{ generated_at }}</p>
</div>

<!-- Student Info -->
<div class="section">
  <div class="section-title">Student Information</div>
  <div class="section-body">
    <table>
      <tr><td><strong>Name</strong></td><td>{{ user.full_name }}</td>
          <td><strong>Email</strong></td><td>{{ user.email }}</td></tr>
      {% if profile %}
      <tr><td><strong>Reg No</strong></td><td>{{ profile.reg_no }}</td>
          <td><strong>Year / Section</strong></td><td>{{ profile.year }} / {{ profile.section or 'N/A' }}</td></tr>
      {% endif %}
    </table>
  </div>
</div>

<!-- LeetCode Stats -->
<div class="section">
  <div class="section-title">LeetCode Statistics</div>
  <div class="section-body">
    {% if lc %}
    <div class="stat-grid">
      <div class="stat-box"><div class="value">{{ lc.total_solved }}</div><div class="label">Total Solved</div></div>
      <div class="stat-box"><div class="value">{{ lc.contest_rating | int }}</div><div class="label">Contest Rating</div></div>
      <div class="stat-box"><div class="value">{{ lc.current_streak }}</div><div class="label">Current Streak (days)</div></div>
      <div class="stat-box"><div class="value" style="color:#22C55E">{{ lc.easy_solved }}</div><div class="label">Easy</div></div>
      <div class="stat-box"><div class="value" style="color:#F59E0B">{{ lc.medium_solved }}</div><div class="label">Medium</div></div>
      <div class="stat-box"><div class="value" style="color:#EF4444">{{ lc.hard_solved }}</div><div class="label">Hard</div></div>
    </div>
    <p style="font-size:10px;color:#64748b">Last synced: {{ lc.last_synced }}</p>
    {% else %}
    <p>No LeetCode data available.</p>
    {% endif %}
  </div>
</div>

<!-- GitHub Stats -->
<div class="section">
  <div class="section-title">GitHub Statistics</div>
  <div class="section-body">
    {% if gh %}
    <div class="stat-grid">
      <div class="stat-box"><div class="value">{{ gh.total_commits }}</div><div class="label">Total Commits</div></div>
      <div class="stat-box"><div class="value">{{ gh.pull_requests }}</div><div class="label">Pull Requests</div></div>
      <div class="stat-box"><div class="value">{{ gh.public_repos }}</div><div class="label">Repositories</div></div>
      <div class="stat-box"><div class="value">{{ gh.stars_received }}</div><div class="label">Stars Received</div></div>
      <div class="stat-box"><div class="value">{{ gh.contribution_streak }}</div><div class="label">Contribution Streak</div></div>
      <div class="stat-box"><div class="value">{{ gh.followers }}</div><div class="label">Followers</div></div>
    </div>
    {% else %}
    <p>No GitHub data available.</p>
    {% endif %}
  </div>
</div>

<!-- Performance Score -->
<div class="section">
  <div class="section-title">Performance Score</div>
  <div class="section-body">
    {% if perf %}
    <div class="stat-grid">
      <div class="stat-box"><div class="value">{{ perf.total_score }}</div><div class="label">Total Score / 100</div></div>
      <div class="stat-box"><div class="value">{{ perf.placement_score }}</div><div class="label">Placement Score</div></div>
      <div class="stat-box"><div class="value" style="color:#22C55E">{{ perf.classification }}</div><div class="label">Classification</div></div>
    </div>
    <div style="margin-top:10px">
      <p><strong>Overall Score</strong></p>
      <div class="score-bar"><div class="score-fill" style="width:{{ perf.total_score }}%"></div></div>
      <p><strong>Placement Readiness</strong></p>
      <div class="score-bar"><div class="score-fill" style="width:{{ perf.placement_score }}%"></div></div>
    </div>
    {% else %}
    <p>Score not yet calculated.</p>
    {% endif %}
  </div>
</div>

<div class="footer">
  <p>{{ college_name }} | CodePulse AI Analytics Platform | Confidential</p>
</div>

</body>
</html>
"""


async def generate_user_pdf_report(
    user,
    profile,
    leetcode_stats,
    github_stats,
    performance,
    ai_analysis: Optional[dict] = None,
) -> bytes:
    """
    Generate a PDF report for a user using WeasyPrint + Jinja2.
    Returns PDF as bytes.
    """
    try:
        from weasyprint import HTML

        env = Environment(loader=BaseLoader())
        template = env.from_string(PDF_TEMPLATE)

        html_content = template.render(
            college_name=settings.COLLEGE_NAME,
            generated_at=datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
            user=user,
            profile=profile,
            lc=leetcode_stats,
            gh=github_stats,
            perf=performance,
            ai=ai_analysis,
        )

        pdf_bytes = HTML(string=html_content).write_pdf()
        return pdf_bytes

    except ImportError:
        logger.warning("WeasyPrint not available, generating minimal PDF")
        return _generate_text_fallback_pdf(user, performance)
    except Exception as exc:
        logger.error(f"PDF generation error: {exc}")
        raise


def _generate_text_fallback_pdf(user, performance) -> bytes:
    """Fallback: generate a simple text-based PDF if WeasyPrint fails."""
    content = f"""CodePulse AI Report
====================
Name: {user.full_name}
Email: {user.email}
Score: {performance.total_score if performance else 'N/A'}
Generated: {datetime.now(tz=timezone.utc).isoformat()}
""".encode("utf-8")
    return content


async def generate_excel_report(db: AsyncSession) -> bytes:
    """
    Generate Excel workbook with Summary, Students, Staff, Daily LeetCode Activity, and Leaderboard sheets.
    Returns bytes.
    """
    import json

    import openpyxl
    from app.models.github import GitHubStats
    from app.models.leetcode import LeetCodeStats
    from app.models.performance import PerformanceScore
    from app.models.profiles import StaffProfile, StudentProfile
    from app.models.user import Role, User
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ─── Style helpers ────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="6366F1")
    header_font = Font(color="FFFFFF", bold=True)

    def write_header(ws, headers: list[str]):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = max(len(h) + 5, 15)

    # ─── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    total_result = await db.execute(select(func.count(User.id)))
    total_users = total_result.scalar() or 0

    active_result = await db.execute(
        select(func.count(User.id)).where(User.is_active == True)  # noqa
    )
    active_users = active_result.scalar() or 0

    avg_result = await db.execute(select(func.avg(PerformanceScore.total_score)))
    avg_score = round(float(avg_result.scalar() or 0), 2)

    write_header(ws_summary, ["Metric", "Value"])
    summary_data = [
        ("Total Users", total_users),
        ("Active Users", active_users),
        ("Average Score", avg_score),
        ("Report Generated", datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
    ]
    for row_idx, (metric, value) in enumerate(summary_data, 2):
        ws_summary.cell(row=row_idx, column=1, value=metric)
        ws_summary.cell(row=row_idx, column=2, value=value)

    # ─── Sheet 2: Students ────────────────────────────────────────────────────
    ws_students = wb.create_sheet("Students")
    student_headers = [
        "Name", "Email", "Reg No", "Year", "Section",
        "LeetCode Username", "GitHub Username",
        "Total Solved", "Easy", "Medium", "Hard",
        "Contest Rating", "Streak",
        "GH Commits", "GH PRs", "GH Repos",
        "Total Score", "LC Score", "GH Score",
        "Placement Score", "Classification", "Placement Class",
    ]
    write_header(ws_students, student_headers)

    students_result = await db.execute(
        select(User, StudentProfile, LeetCodeStats, GitHubStats, PerformanceScore)
        .join(User.role)
        .outerjoin(StudentProfile, StudentProfile.user_id == User.id)
        .outerjoin(LeetCodeStats, LeetCodeStats.user_id == User.id)
        .outerjoin(GitHubStats, GitHubStats.user_id == User.id)
        .outerjoin(PerformanceScore, PerformanceScore.user_id == User.id)
        .where(Role.name == "student")
        .order_by(PerformanceScore.total_score.desc())
    )
    for row_idx, row in enumerate(students_result.fetchall(), 2):
        user, sp, lc, gh, perf = row
        data = [
            user.full_name, user.email,
            sp.reg_no if sp else "", sp.year if sp else "", sp.section if sp else "",
            sp.leetcode_username if sp else "", sp.github_username if sp else "",
            lc.total_solved if lc else 0,
            lc.easy_solved if lc else 0,
            lc.medium_solved if lc else 0,
            lc.hard_solved if lc else 0,
            lc.contest_rating if lc else 0,
            lc.current_streak if lc else 0,
            gh.total_commits if gh else 0,
            gh.pull_requests if gh else 0,
            gh.public_repos if gh else 0,
            perf.total_score if perf else 0,
            perf.leetcode_score if perf else 0,
            perf.github_score if perf else 0,
            perf.placement_score if perf else 0,
            perf.classification if perf else "Needs Improvement",
            perf.placement_classification if perf else "Needs Improvement",
        ]
        for col_idx, val in enumerate(data, 1):
            ws_students.cell(row=row_idx, column=col_idx, value=val)

    # ─── Sheet 3: Staff ───────────────────────────────────────────────────────
    ws_staff = wb.create_sheet("Staff")
    staff_headers = [
        "Name", "Email", "Employee ID", "Designation",
        "LeetCode Username", "GitHub Username",
        "Total Solved", "Easy", "Medium", "Hard",
        "Contest Rating", "Streak",
        "GH Commits", "GH PRs", "GH Repos",
        "Total Score", "LC Score", "GH Score",
        "Classification",
    ]
    write_header(ws_staff, staff_headers)

    staff_result = await db.execute(
        select(User, StaffProfile, LeetCodeStats, GitHubStats, PerformanceScore)
        .join(User.role)
        .outerjoin(StaffProfile, StaffProfile.user_id == User.id)
        .outerjoin(LeetCodeStats, LeetCodeStats.user_id == User.id)
        .outerjoin(GitHubStats, GitHubStats.user_id == User.id)
        .outerjoin(PerformanceScore, PerformanceScore.user_id == User.id)
        .where(Role.name == "staff")
        .order_by(PerformanceScore.total_score.desc())
    )
    for row_idx, row in enumerate(staff_result.fetchall(), 2):
        user, sp, lc, gh, perf = row
        data = [
            user.full_name, user.email,
            sp.employee_id if sp else "", sp.designation if sp else "",
            sp.leetcode_username if sp else "", sp.github_username if sp else "",
            lc.total_solved if lc else 0,
            lc.easy_solved if lc else 0,
            lc.medium_solved if lc else 0,
            lc.hard_solved if lc else 0,
            lc.contest_rating if lc else 0,
            lc.current_streak if lc else 0,
            gh.total_commits if gh else 0,
            gh.pull_requests if gh else 0,
            gh.public_repos if gh else 0,
            perf.total_score if perf else 0,
            perf.leetcode_score if perf else 0,
            perf.github_score if perf else 0,
            perf.classification if perf else "Needs Improvement",
        ]
        for col_idx, val in enumerate(data, 1):
            ws_staff.cell(row=row_idx, column=col_idx, value=val)

    # ─── Sheet 4: Daily Progress ──────────────────────────────────────────────
    ws_daily = wb.create_sheet("Daily Progress")
    daily_headers = ["Name", "Email", "Reg No", "Role", "Date", "LeetCode Problems Solved", "GitHub Commits"]
    write_header(ws_daily, daily_headers)

    users_stats_result = await db.execute(
        select(User, StudentProfile, LeetCodeStats, GitHubStats)
        .outerjoin(StudentProfile, StudentProfile.user_id == User.id)
        .outerjoin(LeetCodeStats, LeetCodeStats.user_id == User.id)
        .outerjoin(GitHubStats, GitHubStats.user_id == User.id)
    )

    daily_progress = {}

    for user, sp, lc, gh in users_stats_result.fetchall():
        reg_no = sp.reg_no if sp else "N/A"
        name = user.full_name
        email = user.email
        role_name = user.role.name

        # 1. LeetCode submissions calendar
        if lc and lc.submission_calendar:
            try:
                lc_cal = json.loads(lc.submission_calendar)
                for ts_str, count in lc_cal.items():
                    if not count or int(count) <= 0:
                        continue
                    ts = int(ts_str)
                    d_str = datetime.fromtimestamp(ts, tz=timezone.utc).date().isoformat()
                    key = (email, d_str)
                    if key not in daily_progress:
                        daily_progress[key] = {
                            "name": name, "email": email, "reg_no": reg_no, "role": role_name, "date": d_str, "leetcode": 0, "github": 0
                        }
                    daily_progress[key]["leetcode"] += int(count)
            except Exception:
                pass

        # 2. GitHub contribution calendar
        if gh and gh.contribution_calendar:
            try:
                gh_cal = json.loads(gh.contribution_calendar)
                for d_str, count in gh_cal.items():
                    if not count or int(count) <= 0:
                        continue
                    key = (email, d_str)
                    if key not in daily_progress:
                        daily_progress[key] = {
                            "name": name, "email": email, "reg_no": reg_no, "role": role_name, "date": d_str, "leetcode": 0, "github": 0
                        }
                    daily_progress[key]["github"] += int(count)
            except Exception:
                pass

    daily_rows = list(daily_progress.values())
    daily_rows.sort(key=lambda x: (x["date"], x["name"]), reverse=True)

    for row_idx, r in enumerate(daily_rows, 2):
        ws_daily.cell(row=row_idx, column=1, value=r["name"])
        ws_daily.cell(row=row_idx, column=2, value=r["email"])
        ws_daily.cell(row=row_idx, column=3, value=r["reg_no"])
        ws_daily.cell(row=row_idx, column=4, value=r["role"])
        ws_daily.cell(row=row_idx, column=5, value=r["date"])
        ws_daily.cell(row=row_idx, column=6, value=r["leetcode"])
        ws_daily.cell(row=row_idx, column=7, value=r["github"])

    # ─── Sheet 5: Leaderboard ─────────────────────────────────────────────────
    ws_leader = wb.create_sheet("Leaderboard")
    write_header(ws_leader, ["Rank", "Name", "Email", "Total Score", "Classification"])

    leader_result = await db.execute(
        select(User, PerformanceScore)
        .join(PerformanceScore, PerformanceScore.user_id == User.id)
        .where(User.is_active == True)  # noqa
        .order_by(PerformanceScore.total_score.desc())
        .limit(50)
    )
    for rank, row in enumerate(leader_result.fetchall(), 1):
        user, perf = row
        data = [rank, user.full_name, user.email, perf.total_score, perf.classification]
        for col_idx, val in enumerate(data, 1):
            ws_leader.cell(row=rank + 1, column=col_idx, value=val)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
